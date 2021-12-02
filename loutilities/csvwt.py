'''
csvwt - write csv from various file types
===================================================
'''

# standard
import pdb
import argparse
import tempfile
import collections
import os
from collections import OrderedDict
from csv import DictReader, DictWriter

# pypi
from xlrd import open_workbook
from openpyxl import load_workbook

# github

# other
from sqlalchemy.orm import class_mapper # see http://www.sqlalchemy.org/ written with 0.8.0b2

# home grown
from . import version

class invalidParameter(Exception): pass
class parameterError(Exception): pass

########################################################################
class _objdict(dict):
########################################################################
    '''
    subclass dict to make it work like an object

    see http://goodcode.io/articles/python-dict-object/
    '''
    def __getattr__(self, name):
        if name in self:
            return self[name]
        else:
            raise AttributeError("No such attribute: " + name)

    def __setattr__(self, name, value):
        self[name] = value

    def __delattr__(self, name):
        if name in self:
            del self[name]
        else:
            raise AttributeError("No such attribute: " + name)

#----------------------------------------------------------------------
def record2csv(inrecs, mapping, outfile=None, encoding=None): 
#----------------------------------------------------------------------
    '''
    convert list of dict or object records to a csv list or file based on a specified mapping
    
    :param inrecs: list of dicts or objects
    :param mapping: OrderedDict {'outfield1':'infield1', 'outfield2':outfunction(inrec), ...} or ['inoutfield1', 'inoutfield2', ...]
    :param outfile: optional output file
    :param encoding: optional file encoding
    :rtype: lines from output file
    '''

    # analyze mapping for outfields
    if isinstance(mapping, list):
        mappingtype = list
    elif type(mapping) in [dict, OrderedDict]:
        mappingtype = dict
    else:
        raise invalidParameter(
            "invalid mapping type {}. mapping type must be list, dict or OrderedDict").with_traceback(
            format(type(mapping)))

    outfields = []
    for outfield in mapping:
        invalue = mapping[outfield] if mappingtype==dict else outfield

        if not isinstance(invalue, str) and not callable(invalue):
            raise invalidParameter('invalid mapping {}. mapping values must be str or function'.format(invalue))

        outfields.append(outfield)

    # create writeable list, csv file
    outreclist = wlist()
    coutreclist = DictWriter(outreclist, outfields)
    coutreclist.writeheader()

    for inrec in inrecs:
        # convert to object if necessary
        if isinstance(inrec, dict):
            inrec = _objdict(inrec)

        outrow = {}
        for outfield in mapping:
            infield = mapping[outfield] if mappingtype==dict else outfield
            if isinstance(infield, str):
                outvalue = getattr(inrec, infield, None)

            else:
                # a function call is requested
                outvalue = infield(inrec)

            outrow[outfield] = outvalue
        
        coutreclist.writerow(outrow)

    # write file if desired
    if outfile:
        with open(outfile, 'w', newline='', encoding=encoding) as out:
            out.writelines(outreclist)

    return outreclist

########################################################################
class Base2Csv():
########################################################################
    '''
    base class for any file to csv conversion
    
    :param outdir: directory to put output file(s) -- if None, temporary directory is used
    '''
    
    #----------------------------------------------------------------------
    def __init__(self, filename, outdir=None, hdrmap=None, encoding=None):
    #----------------------------------------------------------------------
        '''
        '''

        self.filename = filename
        self.tempdir = False
        if outdir is None:
            self.tempdir = True
            outdir = tempfile.mkdtemp(prefix='csvwt-')
        self.dir = outdir
        
        self.files = collections.OrderedDict()
        self.encoding = encoding
        
    #----------------------------------------------------------------------
    def __del__(self):
    #----------------------------------------------------------------------
        '''
        release resources
        '''

        if self.tempdir:
            for name in self.files:
                os.remove(self.files[name])
            os.rmdir(self.dir)
            
    #----------------------------------------------------------------------
    def getfiles(self):
    #----------------------------------------------------------------------
        '''
        get sheetnames and file pathnames produced
        
        :rtype: OrderedDict{sheet:pathname,...}
        '''
        
        return self.files

########################################################################
class Xls2Csv(Base2Csv):
########################################################################
    '''
    create csv file(s) from xlsx sheets
    
    :param filename: name of file to convert
    :param outdir: directory to put output file(s) -- if None, temporary directory is used
    :param hdrmap: maps input header to csv header -- if None, input header is used as csv header
    '''

    def _handle_xlsx(self, filename, hdrmap=None):
        """
        handle xlsx file

        :param filename: xlsx file name
        :param hdrmap: maps input header to csv header -- if None, input header is used as csv header
        """
        # go through each sheet, and save as csv file
        # from http://www.gossamer-threads.com/lists/python/python/833610
        wb = load_workbook(filename) 
        for name in wb.sheetnames: 
            sheet = wb[name] 
            if len(list(sheet.rows)) == 0: continue   # skip empty sheets
            
            # get header
            inhdr = [c.value for c in next(sheet.rows)]
            if hdrmap is not None:
                # NOTE: this has the effect of filtering input columns
                outhdr = [hdrmap[k] for k in hdrmap]
            else:
                hdrmap = dict(list(zip(inhdr,inhdr)))
                outhdr = inhdr
                
            # create output csv file and write header
            self.files[name] = '{0}/{1}.csv'.format(self.dir,name)
            OUT = open(self.files[name], 'w', newline='', encoding=self.encoding)
            writer = DictWriter(OUT,outhdr)
            writer.writeheader()
            
            # copy all the remaining rows in the original sheet to the csv file
            for row in sheet.iter_rows(min_row=1, values_only=True):
                inrow = dict(list(zip(inhdr,row)))
                outrow = {}
                for incol in inhdr:
                    if incol in hdrmap:
                        outrow[hdrmap[incol]] = inrow[incol]
                writer.writerow(outrow)
            
            # we're done with this sheet
            OUT.close()
 
    def _handle_xls(self, filename, hdrmap=None):
        """
        handle xls file

        :param filename: xls file name
        :param hdrmap: maps input header to csv header -- if None, input header is used as csv header
        """
        # go through each sheet, and save as csv file
        # from http://www.gossamer-threads.com/lists/python/python/833610
        wb = open_workbook(filename) 
        for name in wb.sheet_names(): 
            sheet = wb.sheet_by_name(name) 
            if sheet.nrows == 0: continue   # skip empty sheets
            
            # get header
            inhdr = sheet.row_values(0)
            if hdrmap is not None:
                # NOTE: this has the effect of filtering input columns
                outhdr = [hdrmap[k] for k in hdrmap]
            else:
                hdrmap = dict(list(zip(inhdr,inhdr)))
                outhdr = inhdr
                
            # create output csv file and write header
            self.files[name] = '{0}/{1}.csv'.format(self.dir,name)
            OUT = open(self.files[name], 'w', newline='', encoding=self.encoding)
            writer = DictWriter(OUT,outhdr)
            writer.writeheader()
            
            # copy all the rows in the original sheet to the csv file
            for row in range(1,sheet.nrows):
                inrow = dict(list(zip(inhdr,sheet.row_values(row))))
                outrow = {}
                for incol in inhdr:
                    if incol in hdrmap:
                        outrow[hdrmap[incol]] = inrow[incol]
                writer.writerow(outrow)
            
            # we're done with this sheet
            OUT.close()

    def __init__(self, filename, outdir=None, hdrmap=None):
        
        # only handle xlsx these days

        # create outdir if necessary, self.out, self.files
        super().__init__(filename, outdir=outdir)
        
        ext = filename.split('.')[-1]
        if ext.lower() == 'xlsx':
            self._handle_xlsx(filename, hdrmap=hdrmap)
        
        elif ext.lower() == 'xls':
            self._handle_xls(filename, hdrmap=hdrmap)
        
        else:
            raise parameterError(f'invalid extension: {ext}')

########################################################################
class Db2Csv(Base2Csv):
########################################################################
    '''
    create csv file(s) from db tables
        
    :param outdir: directory to put output file(s) -- if None, temporary directory is used
    :param encoding: encoding for csv file, default None (system default)
    '''

    #----------------------------------------------------------------------
    def __init__(self, outdir=None, encoding=None):
    #----------------------------------------------------------------------
        '''
        '''
        
        # create outdir if necessary, self.out, self.files
        super().__init__('', outdir=outdir, encoding=encoding)

        # save encoding for csv file open (is this necessary? we updated Base2Csv to include encoding parameter)
        self.encoding = encoding
        
    #----------------------------------------------------------------------
    def addtable(self, name, session, model, hdrmap=None, **kwargs):
    #----------------------------------------------------------------------
        '''
        insert a new element or update an existing on based on kwargs query
        
        hdrmap may be of the form {infield1:{outfield:function,...},infield2:outfield2,...},
        
            where:
                
                outfield is the column name
                function is defined as function(session,value), session is database session and value is the value of the inrow[infield]
                
                this allows multiple output columns, each tranformed from the input by a different function
        
        :param name: 'sheet' name, used to name output file
        :param session: session within which update occurs
        :param model: table model
        :param hdrmap: maps input table column names to csv header -- if None, input table column names are used as csv header
        :param **kwargs: used for db filter
        '''
        # TODO: currently this only handles flat tables, i.e., if workbook originally was used to make
        # the tables and one sheet made multiple tables, this won't be able to recreate the csv sheet.
        # Not sure what it would take to recreate the csv in that case, but probably has something to do
        # with Mapper.col and Mapper.col.attr -- wish sqlalchemy docs were a bit more clear on this
    
        # get the column names from the model
        inhdr = []
        for col in class_mapper(model).columns:
            inhdr.append(col.key)
        
        # figure out mapping from inhdr to outhdr, and create outhdr
        if hdrmap is not None:
            # NOTE: this has the effect of filtering input columns
            outhndlr = [hdrmap[k] for k in hdrmap]
            outhdr = []
            for k in outhndlr:
                if isinstance(k, str):
                    outhdr.append(k)
                elif isinstance(k, dict):
                    # assumes only one level
                    for subk in k:
                        if not isinstance(subk, str):
                            raise parameterError('{0}: invalid hdrmap {1}'.format(self.filename, hdrmap))
                        outhdr.append(subk)
                else:
                    raise parameterError('{0}: invalid hdrmap {1}'.format(self.filename, hdrmap))
        else:
            hdrmap = dict(list(zip(inhdr,inhdr)))
            outhdr = inhdr

        # create output csv file and write header
        self.files[name] = '{0}/{1}.csv'.format(self.dir,name)
        OUT = open(self.files[name], 'w', newline='', encoding=self.encoding)
        writer = DictWriter(OUT,outhdr)
        writer.writeheader()
            
        # copy all the rows in the table to the csv file
        for inrow in session.query(model).filter_by(**kwargs).all():
            outrow = {}
            for incol in inhdr:
                if incol in hdrmap:
                    outcol = hdrmap[incol]
                    if isinstance(outcol, str):
                        outrow[outcol] = getattr(inrow,incol)
                    # must be dict, call function(session,value) to determine value transformation
                    else:
                        # tranform input for each configured output column
                        for subk in outcol:
                            outrow[subk] = outcol[subk](session,getattr(inrow,incol))
            writer.writerow(outrow)
        
        # we're done with this sheet
        OUT.close()
        
###############################################################################
class wlist(list):
###############################################################################
    '''
    extends list to include a write method, in order to allow csv.writer to output to the list
    '''

    #----------------------------------------------------------------------
    def write(self, line):
    #----------------------------------------------------------------------
        '''
        write method to be added to list object.  Appends line to the list
        
        :param line: line to append to the list
        '''
        self.append(line)   
    
################################################################################
def main():
################################################################################
    '''
    unit tests
    '''

    parser = argparse.ArgumentParser(version='{0} {1}'.format('loutilities',version.__version__))
    parser.add_argument('-d','--dbfilename',help='name of db file for testing')
    #parser.add_argument('--nowuaccess',help='use option to inhibit wunderground access using apikey',action="store_true")
    #parser.add_argument('-l','--loglevel',help='set logging level (default=%(default)s)',default='WARNING')
    #parser.add_argument('-o','--logfile',help='logging output file (default=stdout)',default=sys.stdout)
    args = parser.parse_args()
    
    # act on arguments
    dbfilename = args.dbfilename
    
    # use running.racedb model to test
    from running import racedb
    racedb.setracedb(dbfilename)
    s = racedb.Session()
    
    dd = Db2Csv()

    # note it is ok to split name like this, because it will just get joined together when the csv is processed in clubmember
    hdrmap = {'dateofbirth':'DOB','gender':'Gender',
              'name':{'First':lambda f: ' '.join(f.split(' ')[0:-1]),'Last':lambda f: f.split(' ')[-1]},
              'hometown':{'City':lambda f: ','.join(f.split(',')[0:-1]), 'State': lambda f: f.split(',')[-1]}
                }
    dd.addtable('Sheet1',s,racedb.Runner,hdrmap,active=True)
    files = dd.getfiles()
    
    pdb.set_trace()
    


# ###############################################################################
# ###############################################################################
if __name__ == "__main__":
    main()

