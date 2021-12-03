#!/usr/bin/python
###########################################################################################
#   textreader - read text out of various file types
#
#   Date        Author      Reason
#   ----        ------      ------
#   12/29/12    Lou King    Create
#
#   Copyright 2012 Lou King
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.
#
###########################################################################################
'''
textreader - read text out of various file types
===================================================
'''

# standard
import argparse
import csv

# pypi
from charset_normalizer import detect

# github

# home grown
from . import version

# valid file types
VALIDFTYPES = ['xls','xlsx','docx','txt','csv']
# valid list types
VALIDLTYPES = ['txt','csv']

DOCXTABSIZE = 8
TXTABSIZE = 8

# exceptions for this module.  See __init__.py for package exceptions
class headerError(Exception): pass
class parameterError(Exception): pass

########################################################################
class TextDictReader():
########################################################################
    '''
    get data from a file in various formats
    
    :param filename: name of file to open, or list-like
    :param fieldxform: {'fieldname1':['hdr11, hdr12, ...'], 'fieldname2':[hdr21, ...], ...}, where fieldnamex will be returned in dict, hdrxx are possible headers for columns
    :param reqdfields: list of fields required to be in the header
    :param filetype: if filename is list, this should be filetype list should be interpreted as
    '''
    #----------------------------------------------------------------------
    def __init__(self, filename, fieldxform, reqdfields, filetype='extension'):
    #----------------------------------------------------------------------
        # open the textreader using the file
        self.file = TextReader(filename, filetype)
        self.fieldxform = fieldxform
        self.reqdfields = reqdfields
        
        # self.field item value will be of form {'begin':startindex,'end':startindex+length} for easy slicing
        self.field = {}

        # scan to the header line
        self._findhdr()

    #----------------------------------------------------------------------
    def _findhdr(self):
    #----------------------------------------------------------------------
        '''
        find the header in the file
        '''
    
        foundhdr = False
        delimited = self.file.getdelimited()
        fields = list(self.fieldxform.keys())
        MINMATCHES = len(self.reqdfields)

        # catch StopIteration, which means header wasn't found in the file
        try:
            # loop for each line until header found
            while True:
                origline = next(self.file)
                fieldsfound = 0
                line = []
                if not delimited:
                    for word in origline.split():
                        line.append(word.lower())
                else:
                    for word in origline:
                        line.append(str(word).lower())  # str() called in case non-string returned in origline
                    
                # loop for each potential self.field in a header
                for fieldndx in range(len(fields)):
                    f = fields[fieldndx]
                    match = self.fieldxform[f]
                    
                    # loop for each match possiblity, then within the fields in file header
                    # folding loop likes this gives precedence to the list order of the match possibilities
                    for m in match:
                        # loop for each column in this line, trying to find match possiblities
                        matchfound = False
                        for linendx in range(len(line)):                        
                            # match over the end of the line is no match
                            # m is either a string or a list of strings
                            if isinstance(m, str):
                                m = [m]         # make single string into list
                            if linendx+len(m)>len(line):
                                continue
                            # if we found the match, remember start and end of match
                            if line[linendx:linendx+len(m)] == m:
                                if f not in self.field: self.field[f] = {}
                                self.field[f]['start'] = linendx
                                self.field[f]['end'] = linendx + len(m)
                                self.field[f]['match'] = m
                                self.field[f]['genfield'] = f   # seems redundant, but [f] index is lost later in self.foundfields
                                fieldsfound += 1
                                matchfound = True
                                break   # match possibility loop
                        
                        # found match for this self.field
                        if matchfound: break
                
                # here we've gone through each self.field in the line
                # need to match more than MINMATCHES to call it a header line
                if fieldsfound >= MINMATCHES:
                    
                    # verify that all other required fields are present
                    fieldsnotfound = []
                    for f in self.reqdfields:
                        if f not in self.field:
                            fieldsnotfound.append(f)
                    if len(fieldsnotfound) != 0:
                        raise headerError('could not find fields {} in header {}'.format(fieldsnotfound, origline))
                        
                    # sort found fields by order found within the line
                    foundfields_dec = sorted([(self.field[f]['start'],self.field[f]) for f in self.field])
                    self.foundfields = [ff[1] for ff in foundfields_dec] # get rid of sorting decorator
                        
                    # here we have decided it is a header line
                    # if the file is not delimited, we have to find where these fields start
                    # and tell self.file where the self.field breaks are
                    # assume multi self.field matches are separated by single space
                    if not delimited:
                        # sort found fields by 'start' linendx (self.field number within line)
                        # loop through characters in original line, skipping over spaces within matched fields, to determine
                        # where delimiters should be
                        delimiters = []
                        thischar = 0
                        foundfields_iter = iter(self.foundfields)
                        thisfield = next(foundfields_iter)
                        while True:
                            # scan past the white space
                            while thischar < len(origline) and origline[thischar] == ' ': thischar += 1
                            
                            # we're done looking if we're at the end of the line
                            if thischar == len(origline): break
                            
                            # found a word, remember where it was
                            delimiters.append(thischar)
                            
                            # look for the next match of known header fields
                            matchfound = False
                            if thisfield is not None:
                                # if a match, might be multiple words.  Probably ok to assume single space between them
                                fullmatch = ' '.join(thisfield['match'])
                                if origline[thischar:thischar+len(fullmatch)].lower() == fullmatch:
                                    thischar += len(fullmatch)
                                    matchfound = True
                                    try:
                                        thisfield = next(foundfields_iter)
                                    except StopIteration:
                                        thisfield = None
                            
                            # if found a match, thischar is already updated.  Otherwise, scan past this word
                            if not matchfound:
                                while thischar < len(origline) and origline[thischar] != ' ': thischar += 1
                            
                            # we're done looking if we're at the end of the line
                            if thischar == len(origline): break
                        
                        # set up delimiters in the file reader
                        self.file.setdelimiter(delimiters)
                                    
                    break

            # header fields are in foundfields
            # need to figure out the indeces for data which correspond to the foundfields
            self.fieldhdrs = []
            self.fieldcols = []
            skipped = 0
            for f in self.foundfields:
                self.fieldhdrs.append(f['genfield'])
                currcol = f['start'] - skipped
                self.fieldcols.append(currcol)
                skipped += len(f['match']) - 1  # if matched multiple columns, need to skip some
                
        # not good to come here
        except StopIteration:
            raise headerError('header not found')
        
    #----------------------------------------------------------------------
    def __next__(self):
    #----------------------------------------------------------------------
        '''
        return dict with generic headers and associated data from file
        '''
        
        # get next raw line from the file
        rawline = next(self.file)
        
        # pick columns which are associated with generic headers
        filteredline = [rawline[i] for i in range(len(rawline)) if i in self.fieldcols]
        
        # create dict association, similar to csv.DictReader
        result = dict(list(zip(self.fieldhdrs,filteredline)))
                   
        # and return result
        return result

########################################################################
class TextReader():
########################################################################
    '''
    abstract text reader for several different file types
    creation of object opens TextReader file
    
    :param filename: name of file to open, or list
    '''

    #----------------------------------------------------------------------
    def __init__(self, filename, filetype='extension'):
    #----------------------------------------------------------------------
        """
        open TextReader file
        
        :param filename: name of file to open, or list-like
        :param filetype: if filename is list, this should be filetype list should be interpreted as
        """

        # if true filename, type of filename is string
        if isinstance(filename, str):
            self.ftype = filename.split('.')[-1].lower() # get extension
            self.intype = 'file'
            if self.ftype not in VALIDFTYPES:
                raise parameterError('Invalid filename {}: must have extension in {}'.format(filename, VALIDFTYPES))
        # otherwise assume 'filename' is list-like
        else:
            self.ftype = filetype.lower()
            self.intype = 'list'
            if self.ftype not in VALIDLTYPES:
                raise parameterError('Invalid list: must use filetype in {}'.format(VALIDLTYPES))
        
        # handle excel files
        if self.ftype in ['xls']:
            from xlrd import open_workbook
            self.workbook = open_workbook(filename)
            self.sheet = self.workbook.sheet_by_index(0)    # only first sheet is considered
            self.currrow = 0
            self.nrows = self.sheet.nrows
            self.delimited = True               # rows are already broken into columns
            self.workbook.release_resources()   # sheet is already loaded so we can save memory
            
        # handle excel files
        elif self.ftype in ['xlsx']:
            from openpyxl import load_workbook
            self.workbook = load_workbook(filename)
            self.sheet = self.workbook[self.workbook.sheetnames[0]]    # only first sheet is considered
            self.currrow = 0
            self.rows = list(self.sheet.rows)
            self.nrows = len(list(self.sheet.rows))
            self.delimited = True               # rows are already broken into columns

        # handle word files
        elif self.ftype in ['docx']:
            import docx
            doc = docx.opendocx(filename)
            self.lines = iter(docx.getdocumenttext(doc))
            self.delimited = False
            
        # handle txt files
        elif self.ftype in ['txt']:
            if self.intype == 'file':
                self.TXT = open(filename,'r')
            else:
                self.TXT = iter(filename)
            self.delimited = False
            
        # handle csv files
        elif self.ftype in ['csv']:
            if self.intype == 'file':
                with open(filename, 'rb') as binaryfile:
                    rawdata = binaryfile.read()
                detected = detect(rawdata)
                self._CSV = open(filename, 'r', encoding=detected['encoding'], newline='', errors='replace')
            else:
                self._CSV = iter(filename)
            self.CSV = csv.reader(self._CSV)
            self.delimited = True
            
        self.delimiters = None
        self.opened = True
        
    #----------------------------------------------------------------------
    def __next__(self):
    #----------------------------------------------------------------------
        """
        read next line from TextReader file
        raises StopIteration when end of file reached
        
        :rtype: list of cells for the current row
        """
        
        # check if it is ok to read
        if not self.opened:
            raise ValueError('I/O operation on a closed file')
        
        # handle excel files
        if self.ftype in ['xls']:
            if self.currrow >= self.nrows:
                raise StopIteration
            
            row = self.sheet.row_values(self.currrow)
            self.currrow += 1
            return row
        
        # handle excel files
        if self.ftype in ['xlsx']:
            if self.currrow >= self.nrows:
                raise StopIteration
            
            row = [c.value for c in self.rows[self.currrow]]
            self.currrow += 1
            return row
        
        # handle word files
        elif self.ftype in ['docx']:
            line = next(self.lines)
            line = line.expandtabs(DOCXTABSIZE)
            if self.delimiters:
                splitline = self.delimit(line)
                return splitline
            else:
                return line

        # handle txt files
        elif self.ftype in ['txt']:
            line = next(self.TXT)
            line = line.expandtabs(TXTABSIZE)
            if self.delimiters:
                splitline = self.delimit(line)
                return splitline
            else:
                return line

        # handle txt files
        elif self.ftype in ['csv']:
            row = next(self.CSV)
            return row

    #----------------------------------------------------------------------
    def close(self):
    #----------------------------------------------------------------------
        """
        close TextReader file
        """
        
        self.opened = False
        
        # handle lists
        if self.intype == 'list':
            pass    # no resources to release

        # handle excel files
        elif self.ftype in ['xls','xlsx']:
            pass    # resources have already been released
        
        # handle word files
        elif self.ftype in ['docx']:
            pass    # nothing going on here, either
        
        # handle txt files
        elif self.ftype in ['txt']:
            self.TXT.close()
        
        # handle txt files
        elif self.ftype in ['csv']:
            self._CSV.close()
        

    #----------------------------------------------------------------------
    def getdelimited(self):
    #----------------------------------------------------------------------
        """
        return state whether file is delimited
        
        :rtype: True if delimited
        """
        
        return self.delimited       

    #----------------------------------------------------------------------
    def setdelimiter(self,delimiters):
    #----------------------------------------------------------------------
        """
        set delimiters for file as specified in delimiters
        
        :param delimiters: list of character positions to set delimiters at
        """
        
        # do some validation
        if self.delimited:
            raise parameterError('cannot set delimiters for file which is already delimited')
        if not isinstance(delimiters, list):
            raise parameterError('delimiters must be a list of increasing integers')
        lastdelimiter = -1
        for delimiter in delimiters:
            if not isinstance(delimiter, int) or delimiter <= lastdelimiter:
                raise parameterError('delimiters must be a list of increasing integers')
            lastdelimiter = delimiter
            
        self.delimiters = delimiters
        self.delimited = True
        
    #----------------------------------------------------------------------
    def delimit(self,s):
    #----------------------------------------------------------------------
        """
        split a string based on delimiters
        
        :param s: string to be split
        :rtype: list of string elements, stripped of white space
        """
        
        if not self.delimiters:
            raise parameterError('cannot split string if delimiters not set')
        
        rval = []
        for i in range(len(self.delimiters)):
            start = self.delimiters[i]
            end = self.delimiters[i+1] if i+1 < len(self.delimiters) else None   # last one goes to end of s
            rval.append(s[start:end].strip())
            
        return rval
            
################################################################################
def main():
################################################################################

    parser = argparse.ArgumentParser(version='{0} {1}'.format('loutilities',version.__version__))
    parser.add_argument('filename',help='name of file for testing')
    args = parser.parse_args()
    
    # act on arguments
    filename = args.filename
    
    # open file, print some lines, then close
    ff = TextReader(filename)
    for i in range(6):
        print(next(ff))
    ff.close()

# ###############################################################################
# ###############################################################################
if __name__ == "__main__":
    main()

